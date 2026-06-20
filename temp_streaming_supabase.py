import os
import time
from datetime import datetime
from supabase import create_client, Client
from openpyxl import load_workbook

url: str = "https://seanzwnadqaneusqeami.supabase.co"
key: str = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNlYW56d25hZHFhbmV1c3FlYW1pIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEzNzMxNjIsImV4cCI6MjA5Njk0OTE2Mn0.3-R97YJzSsVW2ecJSW5briUFwFVNAATJHhASB3xgNuI"
supabase: Client = create_client(url, key)

def ingest_data(file_path: str):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Loading {file_path} with openpyxl in read_only mode...")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    
    # 헤더 읽기
    headers = next(rows_iter)
    if headers is None:
        print("Empty file")
        return
        
    header_map = {name: idx for idx, name in enumerate(headers) if name is not None}
    print(f"Headers found: {list(header_map.keys())}")

    required_cols = ['chatId', 'chatDate', 'chatHour', 'userNo', 'userName', 'userDeptName', 'questionTypeCd', 'aiResultStatus', 'userAction', 'prodLvl2Cd', 'prodLvl2Name', 'prodLvl3Cd', 'prodLvl3Name', 'chatMsg', 'userActionValue']

    try:
        print("Clearing existing data...")
        supabase.table("chat_logs").delete().neq("id", 0).execute()
    except Exception as e:
        print(f"Delete warning: {e}")

    chunk_size = 5000
    records = []
    total_inserted = 0

    def extract_hour(time_str):
        if time_str is None:
            return 0
        try:
            return int(str(time_str).split(':')[0])
        except:
            return 0

    print("Iterating rows and uploading in chunks...")
    start_time = time.time()
    
    for row in rows_iter:
        if not any(row):  # 빈 행 무시
            continue
            
        row_dict = {}
        for col in required_cols:
            idx = header_map.get(col)
            val = row[idx] if idx is not None and idx < len(row) else None
            row_dict[col] = val
        
        # chatDate parsing
        chat_date = row_dict.get('chatDate')
        if isinstance(chat_date, datetime):
            chat_date_str = chat_date.strftime('%Y-%m-%d')
        else:
            chat_date_str = str(chat_date)[:10] if chat_date else None
                
        # dept_name
        dept = str(row_dict.get('userDeptName') or '알수없음').replace('서비스센터', '')
        
        record = {
            "chatId": str(row_dict.get('chatId') or ''),
            "chatDate": chat_date_str,
            "chatHour": extract_hour(row_dict.get('chatHour')),
            "userNo": str(row_dict.get('userNo') or '알수없음'),
            "userName": str(row_dict.get('userName') or '알수없음'),
            "userDeptName": dept,
            "questionTypeCd": str(row_dict.get('questionTypeCd') or '기타'),
            "aiResultStatus": str(row_dict.get('aiResultStatus') or '알수없음'),
            "userAction": str(row_dict.get('userAction') or ''),
            "prodLvl2Cd": str(row_dict.get('prodLvl2Cd') or ''),
            "prodLvl2Name": str(row_dict.get('prodLvl2Name') or ''),
            "prodLvl3Cd": str(row_dict.get('prodLvl3Cd') or ''),
            "prodLvl3Name": str(row_dict.get('prodLvl3Name') or ''),
            "chatMsg": str(row_dict.get('chatMsg') or ''),
            "userActionValue": str(row_dict.get('userActionValue') or '')
        }
        records.append(record)
        
        if len(records) >= chunk_size:
            try:
                supabase.table("chat_logs").insert(records).execute()
                total_inserted += len(records)
                print(f"Inserted {total_inserted} records...")
            except Exception as e:
                print(f"Error inserting chunk: {e}")
            records = []
            
    if records:
        try:
            supabase.table("chat_logs").insert(records).execute()
            total_inserted += len(records)
            print(f"Inserted {total_inserted} records...")
        except Exception as e:
            print(f"Error inserting final chunk: {e}")

    wb.close()
    end_time = time.time()
    print(f"Successfully ingested all {total_inserted} records in {round(end_time - start_time, 1)} seconds!")

if __name__ == "__main__":
    target_excel = "RawData_202605_v1.xlsx"
    ingest_data(target_excel)
